# -*-coding:utf-8 -*-
# @author: DanDan
# @contact: 454273687@qq.com
# @file: sheet.py
# @time: 2021/8/29 15:40
# @desc:
from .serializer import SerializerAble
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from .cell import CellSerializer, RowCellsSerializer
from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension
from openpyxl import Workbook
from lxml import etree
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.filters import AutoFilter


class ColumnSerializer(object):
    tag_name = 'Column'

    def __init__(self, index, attrs):
        self.index = index
        self.attrs = attrs


# todo: 完善打印设置

class SheetSerializer(SerializerAble):
    tag_name = 'sheet'

    def __init__(self, title=None):
        self.title = title
        self.rows = []
        self.column_style = []
        self._name = id(self)
        self.freeze_panes = None
        self.auto_filter = None

    @property
    def name(self):
        return self._name

    @name.setter
    def name(self, name):
        if name:
            self._name = name

    @classmethod
    def from_excel(cls, work_sheet: Worksheet, *args, **kwargs):
        self = cls(work_sheet.title)
        # work_sheet.oddHeader
        # work_sheet.print_area
        # work_sheet.print_options
        self.auto_filter = work_sheet.auto_filter
        self.freeze_panes = work_sheet.freeze_panes
        merged_cell_map = {x.start_cell.coordinate: (x.min_row, x.max_row, x.min_col, x.max_col) for x in
                           work_sheet.merged_cells}
        row_dimension_map = {}
        for row_index, row_dimension in work_sheet.row_dimensions.items():
            row_dimension_map[row_index] = row_dimension
        row_index = col_index = 1
        last_row_index = last_col_index = 1
        for column_dimension in work_sheet.column_dimensions.values():
            column_dimension.reindex()
            self.column_style.append(dict(column_dimension))
        for row in work_sheet.iter_rows():
            cells = []
            for cell in row:
                if col_index >= 50:
                    break
                if not isinstance(cell, Cell):
                    col_index += 1
                    continue
                cell_coordinate = cell.coordinate
                row_span = col_span = 0
                if cell_coordinate in merged_cell_map:
                    min_row, max_row, min_col, max_col = merged_cell_map[cell_coordinate]
                    row_span = max_row - min_row
                    col_span = max_col - min_col
                skip_cols = col_index - last_col_index
                cells.append(
                    CellSerializer.from_excel(cell, row_span=row_span, col_span=col_span, skip_cols=skip_cols)
                )
                col_index += 1
                last_col_index = col_index
            if cells:
                skip_rows = row_index - last_row_index
                row_instance = row_dimension_map.get(row_index, None)
                row_serializer = RowCellsSerializer.from_excel(row_instance, index=row_index, cells=cells,
                                                               skip_rows=skip_rows)
                self.rows.append(row_serializer)
            row_index += 1
            if cells:
                last_row_index = row_index
            last_col_index = col_index = 1
        return self

    def to_excel(self, parent: Workbook, *args, **kwargs):
        index = kwargs.get('index')
        sheet = parent.create_sheet(self.title, index=index)
        if self.auto_filter:
            sheet.auto_filter = self.auto_filter
        if self.freeze_panes:
            sheet.freeze_panes = self.freeze_panes
        row_index = 1
        for row in self.rows:
            skip_row = row.to_excel(sheet, row_index=row_index)
            row_index = row_index + skip_row + 1
        for column_style in self.column_style:
            start_column = self.convert_python_value(column_style.get('min', 0), 'int')
            end_column = self.convert_python_value(column_style.get('min', 0), 'max')
            sheet.column_dimensions[get_column_letter(start_column)] = ColumnDimension(
                self, **dict({x: column_style.get(x) for x in ColumnDimension.__fields__}, style=None))

    @classmethod
    def from_xml(cls, node, *args, **kwargs):
        title = cls.convert_python_value(node.attrib.get('title'), 'str')
        self = cls(title)
        self.name = self.convert_python_value(node.attrib.get('name'), 'int')
        style_nodes = kwargs.get('style_node')
        if not len(style_nodes):
            style_nodes = node
        for style_node in style_nodes.xpath(f"//ColumnStyle[@name={self.name}]"):
            self.column_style.append(style_node.attrib)
        freeze_panes_tags = style_nodes.xpath(f'//FreezePanes[@name={self.name}]')
        if freeze_panes_tags:
            freeze_panes_tag = freeze_panes_tags[-1]
            self.freeze_panes = freeze_panes_tag.attrib.get('value')

        auto_filters = style_nodes.xpath(f"//AutoFilters[@name={self.name}]/AutoFilter")
        if auto_filters:
            auto_filter = auto_filters[-1]
            self.auto_filter = AutoFilter.from_tree(auto_filter)
        for row_node in node.xpath(f"{RowCellsSerializer.tag_name}"):
            self.rows.append(RowCellsSerializer.from_xml(row_node, style_node=style_nodes))
        return self

    def to_xml(self, *args, **kwargs):
        name = self.convert_xml_value(self.name)
        sheet_tag = etree.Element(self.tag_name, title=self.convert_xml_value(self.title), name=name)
        style_tags = []
        for column_attr in self.column_style:
            style_tags.append(
                etree.Element('ColumnStyle', name=name, **column_attr)
            )

        if self.freeze_panes:
            style_tags.append(
                etree.Element('FreezePanes', name=name, value=self.freeze_panes)
            )
        if self.auto_filter:
            auto_filters = etree.Element('AutoFilters', name=name)
            auto_filter_tag = self.auto_filter.to_tree('AutoFilter')
            auto_filters.append(auto_filter_tag)
            style_tags.append(auto_filters)
        for row in self.rows:
            row_node, style_nodes = row.to_xml()
            sheet_tag.append(row_node)
            style_tags.extend(style_nodes)
        return sheet_tag, style_tags
