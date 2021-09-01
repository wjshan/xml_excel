# -*-coding:utf-8 -*-
# @author: DanDan
# @contact: 454273687@qq.com
# @file: cell.py
# @time: 2021/8/29 15:21
# @desc:
from .serializer import SerializerAble
from openpyxl.worksheet.dimensions import RowDimension
from openpyxl.cell import Cell
from openpyxl.compat import safe_string
from openpyxl.styles import Font, Border, Fill, Alignment, Protection
from lxml import etree
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import get_column_letter


class RowCellsSerializer(SerializerAble):
    tag_name = "Row"

    def __init__(self, attrs=None, cells=None):
        self.attrs = attrs or {}
        self.cells = cells or []
        self.skip_rows = 0

    @classmethod
    def from_excel(cls, instance: RowDimension, *args, **kwargs):
        cells = kwargs.get('cells', [])
        self = cls(attrs=dict(instance or {}), cells=cells)
        self.skip_rows = kwargs.get('skip_rows')
        return self

    def to_excel(self, parent, *args, **kwargs):
        row_index = kwargs.get('row_index', 1) + self.skip_rows
        parent.row_dimensions[row_index] = RowDimension(parent, **dict(self.attrs, s=None))
        col_index = 1
        for cell in self.cells:
            skip = cell.to_excel(parent, *args, col_index=col_index, row_index=row_index)
            col_index = skip + col_index + 1
        return self.skip_rows

    @classmethod
    def from_xml(cls, node, *args, **kwargs):
        self = cls()
        attrs = dict(node.attrib)
        self.skip_rows = self.convert_python_value(attrs.pop('skip_rows', 0), 'int')
        self.attrs = attrs
        cells = node.xpath(f'{CellSerializer.tag_name}')
        for cell in cells:
            self.cells.append(
                CellSerializer.from_xml(cell, style_node=kwargs.get('style_node'))
            )
        return self

    def to_xml(self, *args, **kwargs):
        row_node = etree.Element(self.tag_name, **self.attrs)
        if self.skip_rows:
            row_node.attrib['skip_rows'] = self.convert_xml_value(self.skip_rows)
        style_nodes = []
        for cell in self.cells:
            cell_node, style_node = cell.to_xml()
            row_node.append(cell_node)
            if len(style_node):
                style_nodes.append(style_node)
        return row_node, style_nodes


class CellSerializer(SerializerAble):
    tag_name = 'Cell'

    style_attrs = {'alignment': Alignment, 'border': Border, 'fill': Fill, 'font': Font, 'protection': Protection}

    def __init__(self):
        self.row_span = 0
        self.col_span = 0
        self.skip_cols = 0
        self.value = None
        self.styles = {}
        self._cell_id = id(self)
        self.number_format = None
        self.clone_num = 0

    @property
    def cell_id(self):
        return self._cell_id

    @cell_id.setter
    def cell_id(self, cell_id):
        if cell_id:
            self._cell_id = cell_id

    @property
    def datatype(self):
        return self.value.__class__.__name__

    @classmethod
    def from_excel(cls, instance: Cell, *args, **kwargs):
        self = cls()
        self.value = instance.value
        self.number_format = instance.number_format
        self.row_span = kwargs.get('row_span', 0)
        self.col_span = kwargs.get('col_span', 0)
        self.skip_cols = kwargs.get('skip_cols', 0)
        for style in self.style_attrs:
            style_instance = getattr(instance, style, None)
            if style_instance:
                self.styles[style] = style_instance
        return self

    def to_excel(self, parent: Worksheet, *args, **kwargs):
        col_index = kwargs.get('col_index', 1) + kwargs.get('skip_cols', self.skip_cols)
        row_index = kwargs.get('row_index', 1)
        column_letter = get_column_letter(col_index)
        cell = parent[f"{column_letter}{row_index}"]
        cell.value = self.value
        # cell = parent.cell(column=col_index, row=row_index, value=self.value)
        if self.number_format:
            cell.number_format = self.number_format
        for style_name, style_instance in self.styles.items():
            setattr(cell, style_name, style_instance)
        min_row = row_index
        max_row = min_row + self.row_span
        min_col = col_index
        max_col = min_col + self.col_span
        if min_row != max_row or min_col != max_col:
            parent.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        clone_num = kwargs.get('clone_num', self.clone_num)
        skip_cols = self.skip_cols
        if clone_num > 1:
            skip_cols += self.to_excel(parent, col_index=col_index + 1, row_index=row_index,
                                       clone_num=clone_num - 1, skip_cols=self.col_span)
        return skip_cols

    @classmethod
    def from_xml(cls, node, *args, **kwargs):
        self = cls()
        style_node = kwargs.get('style_node')
        self.cell_id = self.convert_python_value(node.attrib.get('name'), 'int')
        self.value = self.convert_python_value(node.attrib.get('value', None), node.attrib.get('datatype', 'str'))
        self.row_span = self.convert_python_value(node.attrib.get('row_span', 0), 'int')
        self.col_span = self.convert_python_value(node.attrib.get('col_span', 0), 'int')
        self.skip_cols = self.convert_python_value(node.attrib.get('skip_cols', 0), 'int')
        self.clone_num = self.convert_python_value(node.attrib.get('clone_num', 0), 'int')
        self.number_format = node.attrib.get('number_format')
        if len(style_node):
            styles = style_node.xpath(f'//CellStyle[@name={self.cell_id}]')
        else:
            styles = node.xpath(f'//CellStyle[@name={self.cell_id}]')
        cell_style = styles[-1] if styles else []
        for style in cell_style:
            style_name = style.tag
            self.styles[style_name] = self.style_attrs[style_name].from_tree(style)
        return self

    def to_xml(self, *args, **kwargs):
        value = self.convert_xml_value(self.value)
        name = self.convert_xml_value(self.cell_id)
        cell_node = etree.Element(self.tag_name, name=name, value=value, datatype=str(self.datatype))
        if self.number_format:
            cell_node.attrib['number_format'] = self.number_format
        if self.row_span:
            row_span = self.convert_xml_value(self.row_span)
            cell_node.attrib['row_span'] = row_span
        if self.skip_cols:
            skip_cols = self.convert_xml_value(self.skip_cols)
            cell_node.attrib['skip_cols'] = skip_cols
        if self.col_span:
            col_span = self.convert_xml_value(self.col_span)
            cell_node.attrib['col_span'] = col_span
        if self.clone_num:
            clone_num = self.convert_xml_value(self.clone_num)
            cell_node.attrib['clone_num'] = clone_num
        style_node = etree.Element('CellStyle', name=name)
        for style_name, style_instance in self.styles.items():
            _s = style_instance.to_tree(style_name)
            style_node.append(_s)
        return cell_node, style_node
